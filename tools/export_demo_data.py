"""Export the supplied source workbooks into browser-friendly demo data."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
QUESTIONS_FILE = Path(
    "/Users/sunggat/Downloads/anketa_metaprogrammy_dlya_shkolnikov_36_voprosov (1).xlsx"
)
PROFESSIONS_FILE = Path(
    "/Users/sunggat/Downloads/vesovye_profili_professiy_744 (1).xlsx"
)

POLE_KEYS = {
    "Мотивация К": "motivation_toward",
    "Мотивация ОТ": "motivation_away",
    "Внутренняя референция": "internal_reference",
    "Внешняя референция": "external_reference",
    "Активный": "active",
    "Рефлексивный": "reflective",
    "Возможности": "options",
    "Процедуры": "procedures",
    "Общее": "global",
    "Детали": "detail",
    "Ассоциация": "associated",
    "Диссоциация": "dissociated",
}


def export_questions() -> list[dict]:
    workbook = load_workbook(QUESTIONS_FILE, data_only=True, read_only=True)
    questions_sheet = workbook["Анкета"]
    key_sheet = workbook["Ключ"]
    question_rows = {
        int(row[0]): row
        for row in questions_sheet.iter_rows(min_row=4, max_row=39, values_only=True)
        if row and row[0] is not None
    }
    questions = []
    for row in key_sheet.iter_rows(min_row=4, max_row=39, values_only=True):
        number, scale, yes_pole, no_pole, *_ = row
        qrow = question_rows[int(number)]
        questions.append(
            {
                "id": int(number),
                "scale": scale,
                "text_ru": qrow[2],
                "text_kk": qrow[2],
                "yes_pole": POLE_KEYS[yes_pole],
                "no_pole": POLE_KEYS[no_pole],
            }
        )
    return questions


def export_professions() -> list[dict]:
    workbook = load_workbook(PROFESSIONS_FILE, data_only=True, read_only=True)
    sheet = workbook["Профили профессий"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))]
    professions = []
    for values in sheet.iter_rows(min_row=4, max_row=747, values_only=True):
        if not values or values[0] is None:
            continue
        row = dict(zip(headers, values, strict=False))
        professions.append(
            {
                "id": int(row["ID"]),
                "name_ru": row["Профессия"],
                "name_kk": row["Профессия"],
                "category_ru": row["Профессиональная группа"],
                "category_kk": row["Профессиональная группа"],
                "confidence": float(row["Уверенность профиля"]),
                "confidence_level": row["Уровень уверенности"],
                "meta": {
                    POLE_KEYS[label]: float(row[label])
                    for label in POLE_KEYS
                },
                "radicals": {
                    "paranoid": float(row["Паранойяльный"]),
                    "schizoid": float(row["Шизоидный"]),
                    "epileptoid": float(row["Эпилептоидный"]),
                    "hysteroid": float(row["Истероидный"]),
                    "emotive": float(row["Эмотивный"]),
                    "anxious": float(row["Тревожный"]),
                },
            }
        )
    return professions


def main() -> None:
    out_dir = ROOT / "frontend" / "src" / "data"
    out_dir.mkdir(parents=True, exist_ok=True)
    (out_dir / "questions.json").write_text(
        json.dumps(export_questions(), ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "professions.json").write_text(
        json.dumps(export_professions(), ensure_ascii=False), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
